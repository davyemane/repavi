# reservations/views.py - Système de réservations complet

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg, Case, When, IntegerField
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.core.exceptions import PermissionDenied, ValidationError
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
import csv
import calendar
import traceback

# Imports pour l'export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .models import (
    Reservation, Paiement, TypePaiement, EvaluationReservation, 
    Disponibilite
)
from .forms import (
    ReservationForm, ReservationFilterForm, PaiementForm, EvaluationReservationForm,
    ReponseGestionnaireForm, DisponibiliteForm, AnnulationReservationForm, 
    ModificationReservationForm, StatutReservationForm, RechercheDisponibiliteForm, 
    ExportReservationsForm
)
from home.models import Maison

# Import sécurisé des décorateurs
try:
    from utils.decorators import gestionnaire_required, super_admin_required, client_required
except ImportError:
    def gestionnaire_required(func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, "Vous devez être connecté")
                return redirect('users:login')
            if not hasattr(request.user, 'is_gestionnaire') or not request.user.is_gestionnaire():
                messages.error(request, "Accès réservé aux gestionnaires")
                return redirect('home:index')
            return func(request, *args, **kwargs)
        return wrapper
        
    def super_admin_required(func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, "Vous devez être connecté")
                return redirect('users:login')
            if not hasattr(request.user, 'is_super_admin') or not request.user.is_super_admin():
                messages.error(request, "Accès réservé aux super administrateurs")
                return redirect('home:index')
            return func(request, *args, **kwargs)
        return wrapper
    
    def client_required(func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, "Vous devez être connecté")
                return redirect('users:login')
            if not hasattr(request.user, 'is_client') or not request.user.is_client():
                messages.error(request, "Accès réservé aux clients")
                return redirect('home:index')
            return func(request, *args, **kwargs)
        return wrapper


# ======== VUES PUBLIQUES ET CLIENT ========

def recherche_disponibilite(request):
    """Page de recherche de disponibilité"""
    form = RechercheDisponibiliteForm(request.GET or None)
    maisons_disponibles = []
    
    if form.is_valid():
        # Récupérer les critères
        ville = form.cleaned_data.get('ville')
        categorie = form.cleaned_data.get('categorie')
        date_debut = form.cleaned_data['date_debut']
        date_fin = form.cleaned_data['date_fin']
        nombre_personnes = form.cleaned_data['nombre_personnes']
        prix_min = form.cleaned_data.get('prix_min')
        prix_max = form.cleaned_data.get('prix_max')
        
        # Équipements
        equipements = {}
        for equipement in ['wifi', 'parking', 'piscine', 'climatisation']:
            if form.cleaned_data.get(equipement):
                equipements[equipement] = True
        
        # Filtrer les maisons
        maisons = Maison.objects.filter(
            disponible=True,
            statut_occupation='libre',
            capacite_personnes__gte=nombre_personnes
        ).select_related('ville', 'categorie').prefetch_related('photos')
        
        # Filtres optionnels
        if ville:
            maisons = maisons.filter(ville=ville)
        if categorie:
            maisons = maisons.filter(categorie=categorie)
        if prix_min:
            maisons = maisons.filter(prix_par_nuit__gte=prix_min)
        if prix_max:
            maisons = maisons.filter(prix_par_nuit__lte=prix_max)
        
        # Filtrer par équipements
        for equipement, valeur in equipements.items():
            maisons = maisons.filter(**{equipement: valeur})
        
        # Vérifier la disponibilité pour les dates
        for maison in maisons:
            if Reservation.objects.verifier_disponibilite(maison, date_debut, date_fin):
                # Calculer le prix total
                nombre_nuits = (date_fin - date_debut).days
                prix_total = maison.prix_par_nuit * nombre_nuits
                
                maisons_disponibles.append({
                    'maison': maison,
                    'nombre_nuits': nombre_nuits,
                    'prix_total': prix_total,
                    'prix_par_personne': prix_total / nombre_personnes
                })
    
    context = {
        'form': form,
        'maisons_disponibles': maisons_disponibles,
        'total_resultats': len(maisons_disponibles)
    }
    
    return render(request, 'reservations/recherche_disponibilite.html', context)


@login_required
@gestionnaire_required
def reservations_dashboard(request):
    """Dashboard principal des réservations - ACCESSIBLE AUX GESTIONNAIRES"""
    try:
        # Déterminer les réservations accessibles à l'utilisateur
        if hasattr(request.user, 'is_client') and request.user.is_client():
            reservations_base = Reservation.objects.filter(client=request.user)
        elif hasattr(request.user, 'is_gestionnaire') and request.user.is_gestionnaire():
            reservations_base = Reservation.objects.filter(maison__gestionnaire=request.user)
        elif hasattr(request.user, 'is_super_admin') and request.user.is_super_admin():
            reservations_base = Reservation.objects.all()
        else:
            messages.error(request, "Accès non autorisé.")
            return redirect('home:index')        
        
        # Calcul des statistiques principales - MODIFIÉ POUR ENTIERS
        stats = reservations_base.aggregate(
            total=Count('id'),
            confirmees=Count(Case(When(statut='confirmee', then=1), output_field=IntegerField())),
            en_attente=Count(Case(When(statut='en_attente', then=1), output_field=IntegerField())),
            terminees=Count(Case(When(statut='terminee', then=1), output_field=IntegerField())),
            annulees=Count(Case(When(statut='annulee', then=1), output_field=IntegerField())),
            ca_total=Sum(Case(
                When(statut__in=['confirmee', 'terminee'], then='prix_total'),
                default=0,
                output_field=IntegerField()  # CHANGÉ de DecimalField à IntegerField
            ))
        )
        
        # Calculer les pourcentages
        total = stats['total'] or 1
        stats['pourcentage_confirmees'] = round((stats['confirmees'] / total) * 100, 1)
        stats['pourcentage_en_attente'] = round((stats['en_attente'] / total) * 100, 1)
        
        # S'assurer que ca_total n'est pas None
        stats['ca_total'] = stats['ca_total'] or 0
        
        # Répartition par statut pour le graphique
        repartition_statuts = reservations_base.values('statut').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Évolution mensuelle (6 derniers mois)
        evolution_mensuelle = []
        for i in range(6):
            date_mois = timezone.now().replace(day=1) - timedelta(days=30*i)
            debut_mois = date_mois.replace(day=1)
            fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            count = reservations_base.filter(
                date_creation__gte=debut_mois,
                date_creation__lte=fin_mois
            ).count()
            
            evolution_mensuelle.append({
                'mois': debut_mois.strftime('%m/%Y'),
                'count': count
            })
        
        evolution_mensuelle.reverse()
        
        # Réservations récentes
        reservations_recentes = reservations_base.select_related(
            'maison', 'client'
        ).order_by('-date_creation')[:8]
        
        # Actions requises
        actions_requises = []
        today = timezone.now().date()
        
        # Réservations en attente de confirmation
        en_attente = reservations_base.filter(statut='en_attente').order_by('date_creation')[:5]
        for reservation in en_attente:
            jours_depuis_creation = (today - reservation.date_creation.date()).days
            actions_requises.append({
                'numero': reservation.numero,
                'type_action': 'confirmer',
                'action_description': f'Réservation en attente de confirmation',
                'urgence_jours': jours_depuis_creation,
                'reservation': reservation
            })
        
        # Arrivées prochaines (dans les 7 jours)
        arrivees_prochaines = reservations_base.filter(
            statut='confirmee',
            date_debut__gte=today,
            date_debut__lte=today + timedelta(days=7)
        ).order_by('date_debut')[:3]
        
        for reservation in arrivees_prochaines:
            jours_jusqu_arrivee = (reservation.date_debut - today).days
            actions_requises.append({
                'numero': reservation.numero,
                'type_action': 'arrivee',
                'action_description': f'Arrivée prévue le {reservation.date_debut.strftime("%d/%m")}',
                'urgence_jours': jours_jusqu_arrivee,
                'reservation': reservation
            })
        
        # Trier les actions par urgence
        actions_requises.sort(key=lambda x: x['urgence_jours'])
        actions_requises = actions_requises[:8]  # Limiter à 8 éléments
        
        # Métriques de performance (pour gestionnaires)
        metriques = {}
        if request.user.is_gestionnaire() or request.user.is_super_admin():
            # Taux d'occupation ce mois
            debut_mois = today.replace(day=1)
            fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            jours_mois = (fin_mois - debut_mois).days + 1
            
            # Calculer les nuits occupées ce mois
            reservations_mois = reservations_base.filter(
                statut__in=['confirmee', 'terminee'],
                date_debut__lte=fin_mois,
                date_fin__gte=debut_mois
            )
            
            nuits_occupees = 0
            for reservation in reservations_mois:
                debut_periode = max(reservation.date_debut, debut_mois)
                fin_periode = min(reservation.date_fin, fin_mois)
                if debut_periode <= fin_periode:
                    nuits_occupees += (fin_periode - debut_periode).days
            
            # Nombre total de maisons gérées
            if request.user.is_gestionnaire():
                nb_maisons = Maison.objects.filter(gestionnaire=request.user).count()
            else:
                nb_maisons = Maison.objects.count()
            
            # Calcul du taux d'occupation
            nuits_possibles = nb_maisons * jours_mois
            taux_occupation = round((nuits_occupees / nuits_possibles) * 100, 1) if nuits_possibles > 0 else 0
            
            # Durée moyenne de séjour (6 derniers mois)
            six_mois_ago = today - timedelta(days=180)
            reservations_recentes_stats = reservations_base.filter(
                statut__in=['confirmee', 'terminee'],
                date_creation__gte=six_mois_ago
            )
            
            duree_moyenne = reservations_recentes_stats.aggregate(
                moyenne=Avg('nombre_nuits')
            )['moyenne'] or 0
            
            # Panier moyen
            panier_moyen = reservations_recentes_stats.aggregate(
                moyenne=Avg('prix_total')
            )['moyenne'] or 0
            
            # Taux d'annulation
            total_reservations_periode = reservations_base.filter(date_creation__gte=six_mois_ago).count()
            annulations_periode = reservations_base.filter(
                statut='annulee',
                date_creation__gte=six_mois_ago
            ).count()
            
            taux_annulation = round((annulations_periode / total_reservations_periode) * 100, 1) if total_reservations_periode > 0 else 0
            
            metriques = {
                'taux_occupation': taux_occupation,
                'duree_moyenne_sejour': duree_moyenne,
                'panier_moyen': panier_moyen,
                'taux_annulation': taux_annulation
            }
        
        # Prochaines échéances
        prochaines_echeances = []
        
        # Arrivées dans les 3 prochains jours
        arrivees = reservations_base.filter(
            statut='confirmee',
            date_debut__gte=today,
            date_debut__lte=today + timedelta(days=3)
        ).order_by('date_debut')[:3]
        
        for reservation in arrivees:
            prochaines_echeances.append({
                'type': 'arrivee',
                'date': reservation.date_debut,
                'reservation': reservation
            })
        
        # Départs dans les 3 prochains jours
        departs = reservations_base.filter(
            statut='confirmee',
            date_fin__gte=today,
            date_fin__lte=today + timedelta(days=3)
        ).order_by('date_fin')[:3]
        
        for reservation in departs:
            prochaines_echeances.append({
                'type': 'depart',
                'date': reservation.date_fin,
                'reservation': reservation
            })
        
        # Trier par date
        prochaines_echeances.sort(key=lambda x: x['date'])
        prochaines_echeances = prochaines_echeances[:6]  # Limiter à 6
        
        context = {
            'stats': stats,
            'repartition_statuts': repartition_statuts,
            'evolution_mensuelle': evolution_mensuelle,
            'reservations_recentes': reservations_recentes,
            'actions_requises': actions_requises,
            'metriques': metriques,
            'prochaines_echeances': prochaines_echeances,
            'user_type': 'client' if request.user.is_client() else 'gestionnaire'
        }
        
        return render(request, 'reservations/dashboard.html', context)
        
    except Exception as e:
        # En cas d'erreur, afficher un dashboard minimal
        messages.error(request, f"Erreur lors du chargement du dashboard: {str(e)}")
        
        context = {
            'stats': {
                'total': 0, 'confirmees': 0, 'en_attente': 0, 'terminees': 0, 'annulees': 0,
                'ca_total': 0, 'pourcentage_confirmees': 0, 'pourcentage_en_attente': 0
            },
            'repartition_statuts': [],
            'evolution_mensuelle': [],
            'reservations_recentes': [],
            'actions_requises': [],
            'metriques': {},
            'prochaines_echeances': [],
            'user_type': 'client' if request.user.is_client() else 'gestionnaire'
        }
        
        return render(request, 'reservations/dashboard.html', context)


@login_required
@client_required
def reserver_maison(request, maison_slug):
    """Page de réservation d'une maison"""
    maison = get_object_or_404(
        Maison.objects.select_related('ville', 'categorie'),
        slug=maison_slug,
        disponible=True
    )
    
    # Vérifier les paramètres de recherche
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    nombre_personnes = request.GET.get('nombre_personnes', 2)
    
    initial_data = {'nombre_personnes': nombre_personnes}
    if date_debut:
        try:
            initial_data['date_debut'] = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_fin:
        try:
            initial_data['date_fin'] = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.method == 'POST':
        form = ReservationForm(request.POST, user=request.user, maison=maison)
        if form.is_valid():
            try:
                with transaction.atomic():
                    reservation = form.save()
                    
                    # Messages de succès dans le bloc atomic
                    messages.success(
                        request, 
                        f'Votre réservation {reservation.numero} a été créée avec succès! '
                        'Vous allez recevoir un email de confirmation.'
                    )
                
                # Redirection en dehors du bloc atomic
                return redirect('reservations:detail', numero=reservation.numero)
                
            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Erreur lors de la création de la réservation: {str(e)}')
                print(f"Erreur complète: {traceback.format_exc()}")
    else:
        form = ReservationForm(initial=initial_data, user=request.user, maison=maison)
    
    # Calculer le prix estimé si les dates sont fournies
    prix_estime = None
    if date_debut and date_fin:
        try:
            d_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            d_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
            nombre_nuits = (d_fin - d_debut).days
            if nombre_nuits > 0:
                prix_estime = {
                    'nombre_nuits': nombre_nuits,
                    'prix_par_nuit': maison.prix_par_nuit,
                    'sous_total': maison.prix_par_nuit * nombre_nuits,
                    'frais_service': 0,  # TODO: calculer les frais
                    'total': maison.prix_par_nuit * nombre_nuits
                }
        except ValueError:
            pass
    
    # Import sécurisé pour les types de paiement
    try:
        types_paiement = TypePaiement.objects.filter(actif=True)
    except:
        types_paiement = []
    
    # CORRECTION: Récupération des évaluations (remplacer Avis par EvaluationReservation)
    try:
        # Utiliser EvaluationReservation au lieu d'Avis non défini
        evaluations = EvaluationReservation.objects.filter(
            reservation__maison=maison,
            approuve=True
        ).select_related('reservation__client').order_by('-date_creation')[:10]
    except Exception as e:
        print(f"Erreur lors de la récupération des évaluations: {e}")
        evaluations = []
    
    # Calcul de la note moyenne et du nombre d'évaluations
    try:
        if evaluations:
            note_moyenne = sum(eval.note_globale for eval in evaluations) / len(evaluations)
            note_moyenne = round(note_moyenne, 1)
            nombre_evaluations = len(evaluations)
        else:
            note_moyenne = 0
            nombre_evaluations = 0
    except:
        note_moyenne = 0
        nombre_evaluations = 0
    
    context = {
        'maison': maison,
        'form': form,
        'prix_estime': prix_estime,
        'types_paiement': types_paiement,
        'evaluations': evaluations,  # CHANGÉ de 'avis' à 'evaluations'
        'note_moyenne': note_moyenne,
        'nombre_evaluations': nombre_evaluations,  # CHANGÉ de 'nombre_avis'
    }
    
    return render(request, 'reservations/reserver_maison.html', context)


@login_required
def mes_reservations(request):
    """Liste des réservations de l'utilisateur connecté"""
    if request.user.is_client():
        reservations = Reservation.objects.filter(client=request.user)
    elif request.user.is_gestionnaire():
        reservations = Reservation.objects.filter(maison__gestionnaire=request.user)
    elif request.user.is_super_admin():
        reservations = Reservation.objects.all()
    else:
        messages.error(request, "Accès non autorisé.")
        return redirect('home:index')
    
    # Filtres
    form = ReservationFilterForm(request.GET, user=request.user)
    
    if form.is_valid():
        search = form.cleaned_data.get('search')
        statut = form.cleaned_data.get('statut')
        maison = form.cleaned_data.get('maison')
        date_debut = form.cleaned_data.get('date_debut')
        date_fin = form.cleaned_data.get('date_fin')
        mode_paiement = form.cleaned_data.get('mode_paiement')
        
        if search:
            reservations = reservations.filter(
                Q(numero__icontains=search) |
                Q(client__first_name__icontains=search) |
                Q(client__last_name__icontains=search) |
                Q(maison__nom__icontains=search) |
                Q(maison__numero__icontains=search)
            )
        
        if statut:
            reservations = reservations.filter(statut=statut)
        if maison:
            reservations = reservations.filter(maison=maison)
        if date_debut:
            reservations = reservations.filter(date_debut__gte=date_debut)
        if date_fin:
            reservations = reservations.filter(date_fin__lte=date_fin)
        if mode_paiement:
            reservations = reservations.filter(mode_paiement=mode_paiement)
    
    # Tri
    sort_by = request.GET.get('sort', '-date_creation')
    valid_sorts = ['-date_creation', 'date_creation', 'date_debut', '-date_debut', 'statut', 'prix_total']
    if sort_by in valid_sorts:
        reservations = reservations.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(reservations.select_related('maison', 'client'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques pour l'utilisateur
    stats = {
        'total': reservations.count(),
        'en_attente': reservations.filter(statut='en_attente').count(),
        'confirmees': reservations.filter(statut='confirmee').count(),
        'terminees': reservations.filter(statut='terminee').count(),
        'annulees': reservations.filter(statut='annulee').count(),
    }
    
    if request.user.is_client():
        stats['montant_total'] = reservations.filter(
            statut__in=['confirmee', 'terminee']
        ).aggregate(total=Sum('prix_total'))['total'] or 0
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'sort_by': sort_by,
        'stats': stats,
        'user_type': 'client' if request.user.is_client() else 'gestionnaire'
    }
    
    return render(request, 'reservations/mes_reservations.html', context)


@login_required
def detail_reservation(request, numero):
    """Détail d'une réservation"""
    reservation = get_object_or_404(
        Reservation.objects.select_related('maison', 'client', 'annulee_par')
                           .prefetch_related('paiements', 'maison__photos'),
        numero=numero
    )
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:mes_reservations')
    
    # Récupérer les paiements
    paiements = reservation.paiements.all().order_by('date_creation')
    
    # Récupérer l'évaluation si elle existe
    evaluation = None
    try:
        evaluation = reservation.evaluation
    except EvaluationReservation.DoesNotExist:
        pass
    
    # Actions possibles selon l'utilisateur et l'état
    actions_possibles = {
        'peut_modifier': reservation.est_modifiable and (
            request.user == reservation.client or 
            request.user == reservation.maison.gestionnaire or
            request.user.is_super_admin()
        ),
        'peut_annuler': reservation.est_annulable and (
            request.user == reservation.client or 
            request.user == reservation.maison.gestionnaire or
            request.user.is_super_admin()
        ),
        'peut_confirmer': (
            reservation.statut == 'en_attente' and
            (request.user == reservation.maison.gestionnaire or request.user.is_super_admin())
        ),
        'peut_terminer': (
            reservation.statut == 'confirmee' and
            reservation.date_fin <= timezone.now().date() and
            (request.user == reservation.maison.gestionnaire or request.user.is_super_admin())
        ),
        'peut_evaluer': (
            reservation.statut == 'terminee' and
            request.user == reservation.client and
            not evaluation
        ),
        'peut_payer': (
            reservation.statut == 'confirmee' and
            request.user == reservation.client and
            reservation.montant_restant > 0
        )
    }
    
    context = {
        'reservation': reservation,
        'paiements': paiements,
        'evaluation': evaluation,
        'actions_possibles': actions_possibles,
        'montant_paye': paiements.filter(statut='valide').aggregate(
            total=Sum('montant')
        )['total'] or 0
    }
    
    return render(request, 'reservations/detail_reservation.html', context)


@login_required
def modifier_reservation(request, numero):
    """Modifier une réservation"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:mes_reservations')
    
    if not reservation.est_modifiable:
        messages.error(request, "Cette réservation ne peut plus être modifiée.")
        return redirect('reservations:detail', numero=numero)
    
    if request.method == 'POST':
        form = ModificationReservationForm(request.POST, instance=reservation)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    messages.success(request, 'Réservation modifiée avec succès.')
                    
                    # TODO: Envoyer notification de modification
                    
                    return redirect('reservations:detail', numero=numero)
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification: {str(e)}')
    else:
        form = ModificationReservationForm(instance=reservation)
    
    context = {
        'reservation': reservation,
        'form': form
    }
    
    return render(request, 'reservations/modifier_reservation.html', context)


@login_required
def annuler_reservation(request, numero):
    """Annuler une réservation"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:mes_reservations')
    
    if not reservation.est_annulable:
        messages.error(request, "Cette réservation ne peut plus être annulée.")
        return redirect('reservations:detail', numero=numero)
    
    if request.method == 'POST':
        form = AnnulationReservationForm(request.POST)
        if form.is_valid():
            try:
                raison_complete = form.cleaned_data['raison']
                commentaire = form.cleaned_data.get('commentaire', '')
                if commentaire:
                    raison_complete += f" - {commentaire}"
                
                reservation.annuler(raison_complete, request.user)
                
                messages.success(request, 'Réservation annulée avec succès.')
                
                # TODO: Envoyer notification d'annulation
                # TODO: Gérer les remboursements si nécessaire
                
                return redirect('reservations:detail', numero=numero)
            except ValidationError as e:
                messages.error(request, str(e))
    else:
        form = AnnulationReservationForm()
    
    context = {
        'reservation': reservation,
        'form': form
    }
    
    return render(request, 'reservations/annuler_reservation.html', context)


# ======== VUES GESTIONNAIRE ========

@login_required
@gestionnaire_required
def tableau_bord_reservations(request):
    """Tableau de bord des réservations pour gestionnaires"""
    # Statistiques générales
    reservations_base = Reservation.objects.filter(maison__gestionnaire=request.user)
    
    stats = {
        'total_reservations': reservations_base.count(),
        'en_attente': reservations_base.filter(statut='en_attente').count(),
        'confirmees': reservations_base.filter(statut='confirmee').count(),
        'terminees_mois': reservations_base.filter(
            statut='terminee',
            date_fin__gte=timezone.now().replace(day=1).date()
        ).count(),
        'ca_mois': reservations_base.filter(
            statut__in=['confirmee', 'terminee'],
            date_creation__gte=timezone.now().replace(day=1)
        ).aggregate(total=Sum('prix_total'))['total'] or 0,
    }
    
    # Réservations nécessitant une action
    actions_requises = {
        'en_attente': reservations_base.filter(statut='en_attente').order_by('date_creation')[:5],
        'arrivees_aujourdhui': reservations_base.filter(
            statut='confirmee',
            date_debut=timezone.now().date()
        ).order_by('heure_arrivee'),
        'departs_aujourdhui': reservations_base.filter(
            statut='confirmee',
            date_fin=timezone.now().date()
        ).order_by('heure_depart'),
    }
    
    # Dernières réservations
    dernieres_reservations = reservations_base.select_related(
        'client', 'maison'
    ).order_by('-date_creation')[:10]
    
    # Graphique des réservations par mois (6 derniers mois)
    mois_stats = []
    for i in range(6):
        date_mois = timezone.now().replace(day=1) - timedelta(days=30*i)
        debut_mois = date_mois.replace(day=1)
        fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = reservations_base.filter(
            date_creation__gte=debut_mois,
            date_creation__lte=fin_mois
        ).count()
        
        mois_stats.append({
            'mois': debut_mois.strftime('%m/%Y'),
            'count': count
        })
    
    mois_stats.reverse()
    
    context = {
        'stats': stats,
        'actions_requises': actions_requises,
        'dernieres_reservations': dernieres_reservations,
        'mois_stats': mois_stats,
    }
    
    return render(request, 'reservations/tableau_bord_gestionnaire.html', context)


@login_required
@gestionnaire_required
def gerer_reservation(request, numero):
    """Gérer une réservation (gestionnaire) - AMÉLIORÉ"""
    reservation = get_object_or_404(
        Reservation.objects.select_related('maison', 'client'),
        numero=numero
    )
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:dashboard')
    
    if request.method == 'POST':
        nouveau_statut = request.POST.get('nouveau_statut')
        commentaire = request.POST.get('commentaire', '')
        
        try:
            with transaction.atomic():
                ancien_statut = reservation.statut
                
                if nouveau_statut == 'confirmee':
                    reservation.confirmer(request.user)
                    # OCCUPER LA MAISON
                    if reservation.maison:
                        reservation.maison.occuper_maison(
                            locataire=reservation.client,
                            date_fin=reservation.date_fin
                        )
                    messages.success(request, f'Réservation confirmée. Maison occupée par {reservation.client.get_full_name()}.')
                    
                elif nouveau_statut == 'terminee':
                    reservation.terminer()
                    # LIBÉRER LA MAISON
                    if reservation.maison:
                        reservation.maison.liberer_maison()
                    messages.success(request, f'Réservation terminée. Maison libérée.')
                    
                elif nouveau_statut == 'annulee':
                    reservation.annuler(commentaire or "Annulée par le gestionnaire", request.user)
                    # LIBÉRER LA MAISON SI ELLE ÉTAIT OCCUPÉE
                    if reservation.maison and reservation.maison.locataire_actuel == reservation.client:
                        reservation.maison.liberer_maison()
                    messages.success(request, f'Réservation annulée. Maison libérée.')
                
                if commentaire:
                    reservation.commentaire_gestionnaire = commentaire
                    reservation.save()
                
                return redirect('reservations:detail', numero=numero)
                
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement: {str(e)}')
    
    # Afficher le formulaire
    form = StatutReservationForm(reservation=reservation)
    
    context = {
        'reservation': reservation,
        'form': form
    }
    
    return render(request, 'reservations/gerer_reservation.html', context)


# ======== VUES PAIEMENTS ========

@login_required
def paiements_reservation(request, numero):
    """Gestion des paiements d'une réservation"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:mes_reservations')
    
    paiements = reservation.paiements.all().order_by('date_creation')
    
    # Calculer les totaux
    montant_paye = paiements.filter(statut='valide').aggregate(
        total=Sum('montant')
    )['total'] or 0
    
    montant_restant = reservation.prix_total - montant_paye
    
    context = {
        'reservation': reservation,
        'paiements': paiements,
        'montant_paye': montant_paye,
        'montant_restant': montant_restant,
        'peut_ajouter_paiement': (
            request.user == reservation.client or 
            request.user == reservation.maison.gestionnaire or
            request.user.is_super_admin()
        ) and montant_restant > 0
    }
    
    return render(request, 'reservations/paiements_reservation.html', context)


@login_required
def ajouter_paiement(request, numero):
    """Ajouter un paiement à une réservation"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:mes_reservations')
    
    if request.method == 'POST':
        form = PaiementForm(request.POST, reservation=reservation)
        if form.is_valid():
            paiement = form.save()
            
            # Si c'est un gestionnaire ou admin, valider automatiquement
            if request.user.is_gestionnaire() or request.user.is_super_admin():
                paiement.valider(notes="Validé par le gestionnaire")
            
            messages.success(request, 'Paiement ajouté avec succès.')
            return redirect('reservations:paiements', numero=numero)
    else:
        form = PaiementForm(reservation=reservation)
    
    context = {
        'reservation': reservation,
        'form': form
    }
    
    return render(request, 'reservations/ajouter_paiement.html', context)


@login_required
@gestionnaire_required
def dashboard_paiements(request):
    """Vue dédiée aux paiements depuis le dashboard"""
    # Récupérer les paiements selon les permissions
    if hasattr(request.user, 'is_super_admin') and request.user.is_super_admin():
        paiements = Paiement.objects.all()
    elif hasattr(request.user, 'is_gestionnaire') and request.user.is_gestionnaire():
        paiements = Paiement.objects.filter(reservation__maison__gestionnaire=request.user)
    else:
        paiements = Paiement.objects.filter(reservation__client=request.user)
    
    # Paiements en attente
    paiements_en_attente = paiements.filter(statut='en_attente').select_related(
        'reservation__maison', 'reservation__client', 'type_paiement'
    ).order_by('-date_creation')
    
    # Paiements récents
    paiements_recents = paiements.filter(statut='valide').select_related(
        'reservation__maison', 'reservation__client', 'type_paiement'
    ).order_by('-date_validation')[:10]
    
    context = {
        'paiements_en_attente': paiements_en_attente,
        'paiements_recents': paiements_recents,
        'stats': {
            'total_paiements': paiements.count(),
            'en_attente': paiements_en_attente.count(),
            'valides': paiements.filter(statut='valide').count(),
            'echecs': paiements.filter(statut='echec').count(),
        }
    }
    
    return render(request, 'reservations/dashboard_paiements.html', context)


@login_required
@gestionnaire_required
@require_http_methods(["POST"])
def valider_paiement(request, paiement_id):
    """Valider un paiement depuis le dashboard"""
    paiement = get_object_or_404(Paiement, id=paiement_id)
    
    # Vérifier les permissions
    if not paiement.reservation.can_be_managed_by(request.user):
        return JsonResponse({'success': False, 'error': 'Permission refusée'})
    
    if paiement.statut != 'en_attente':
        return JsonResponse({'success': False, 'error': 'Ce paiement ne peut pas être validé'})
    
    try:
        notes = request.POST.get('notes', f'Validé par {request.user.get_full_name()}')
        paiement.valider(notes=notes)
        
        return JsonResponse({
            'success': True,
            'message': f'Paiement de {paiement.montant:,} FCFA validé avec succès'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ======== VUES ÉVALUATIONS ========

@login_required
@client_required
def evaluer_reservation(request, numero):
    """Évaluer une réservation terminée"""
    reservation = get_object_or_404(
        Reservation,
        numero=numero,
        client=request.user,
        statut='terminee'
    )
    
    # Vérifier qu'il n'y a pas déjà une évaluation
    try:
        evaluation = reservation.evaluation
        messages.info(request, "Vous avez déjà évalué cette réservation.")
        return redirect('reservations:detail', numero=numero)
    except EvaluationReservation.DoesNotExist:
        pass
    
    if request.method == 'POST':
        form = EvaluationReservationForm(request.POST)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.reservation = reservation
            evaluation.save()
            
            messages.success(request, 'Merci pour votre évaluation!')
            
            # TODO: Notifier le gestionnaire
            
            return redirect('reservations:detail', numero=numero)
    else:
        form = EvaluationReservationForm()
    
    context = {
        'reservation': reservation,
        'form': form
    }
    
    return render(request, 'reservations/evaluer_reservation.html', context)


@login_required
@gestionnaire_required
def repondre_evaluation(request, numero):
    """Répondre à une évaluation"""
    reservation = get_object_or_404(
        Reservation,
        numero=numero,
        maison__gestionnaire=request.user
    )
    
    try:
        evaluation = reservation.evaluation
    except EvaluationReservation.DoesNotExist:
        messages.error(request, "Aucune évaluation trouvée pour cette réservation.")
        return redirect('reservations:detail', numero=numero)
    
    if request.method == 'POST':
        form = ReponseGestionnaireForm(request.POST, instance=evaluation)
        if form.is_valid():
            form.save()
            messages.success(request, 'Votre réponse a été ajoutée.')
            return redirect('reservations:detail', numero=numero)
    else:
        form = ReponseGestionnaireForm(instance=evaluation)
    
    context = {
        'reservation': reservation,
        'evaluation': evaluation,
        'form': form
    }
    
    return render(request, 'reservations/repondre_evaluation.html', context)


@login_required
@gestionnaire_required
def valider_reservation(request, numero):
    """Valider une réservation directement depuis le dashboard"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:dashboard')
    
    if reservation.statut != 'en_attente':
        messages.error(request, "Seules les réservations en attente peuvent être validées.")
        return redirect('reservations:detail', numero=numero)
    
    try:
        with transaction.atomic():
            # Confirmer la réservation
            reservation.confirmer(request.user)
            
            # NOUVELLE LOGIQUE: Occuper automatiquement la maison
            if reservation.maison:
                reservation.maison.occuper_maison(
                    locataire=reservation.client,
                    date_fin=reservation.date_fin
                )
            
            messages.success(request, f'Réservation {reservation.numero} validée avec succès. La maison est maintenant occupée.')
            
            # TODO: Envoyer notification au client
            
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Erreur lors de la validation: {str(e)}')
    
    return redirect('reservations:detail', numero=numero)


@login_required
@gestionnaire_required
def terminer_reservation(request, numero):
    """Terminer une réservation et libérer la maison"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    # Vérifier les permissions
    if not reservation.can_be_managed_by(request.user):
        messages.error(request, "Vous n'avez pas accès à cette réservation.")
        return redirect('reservations:dashboard')
    
    if reservation.statut != 'confirmee':
        messages.error(request, "Seules les réservations confirmées peuvent être terminées.")
        return redirect('reservations:detail', numero=numero)
    
    try:
        with transaction.atomic():
            # Terminer la réservation
            reservation.terminer()
            
            # NOUVELLE LOGIQUE: Libérer automatiquement la maison
            if reservation.maison:
                reservation.maison.liberer_maison()
            
            messages.success(request, f'Réservation {reservation.numero} terminée avec succès. La maison est maintenant libre.')
            
            # TODO: Demander évaluation au client
            
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Erreur lors de la terminaison: {str(e)}')
    
    return redirect('reservations:detail', numero=numero)


# ======== VUES CALENDRIER ET DISPONIBILITÉS ========

@login_required
@gestionnaire_required
def calendrier_reservations(request):
    """Calendrier des réservations"""
    # Paramètres de date
    today = timezone.now().date()
    mois = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    
    # Créer les dates du mois
    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])
    
    # Récupérer les réservations du mois
    reservations = Reservation.objects.filter(
        maison__gestionnaire=request.user,
        date_debut__lte=dernier_jour,
        date_fin__gte=premier_jour
    ).exclude(statut='annulee').select_related('maison', 'client')
    
    # Organiser par date
    reservations_par_date = {}
    date_courante = premier_jour
    while date_courante <= dernier_jour:
        reservations_jour = []
        for reservation in reservations:
            if reservation.date_debut <= date_courante <= reservation.date_fin:
                reservations_jour.append(reservation)
        reservations_par_date[date_courante] = reservations_jour
        date_courante += timedelta(days=1)
    
    # Navigation
    mois_precedent = premier_jour - timedelta(days=1)
    mois_suivant = dernier_jour + timedelta(days=1)
    
    context = {
        'mois': mois,
        'annee': annee,
        'nom_mois': calendar.month_name[mois],
        'premier_jour': premier_jour,
        'dernier_jour': dernier_jour,
        'reservations_par_date': reservations_par_date,
        'mois_precedent': mois_precedent,
        'mois_suivant': mois_suivant,
        'jours_semaine': ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim'],
    }
    
    return render(request, 'reservations/calendrier_reservations.html', context)


@login_required
@gestionnaire_required
def gerer_disponibilites(request, maison_id):
    """Gérer les disponibilités d'une maison"""
    maison = get_object_or_404(
        Maison,
        id=maison_id,
        gestionnaire=request.user
    )
    
    # Paramètres de date
    today = timezone.now().date()
    mois = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    
    premier_jour = date(annee, mois, 1)
    dernier_jour = date(annee, mois, calendar.monthrange(annee, mois)[1])
    
    # Récupérer les disponibilités existantes
    disponibilites = Disponibilite.objects.filter(
        maison=maison,
        date__gte=premier_jour,
        date__lte=dernier_jour
    )
    
    # Créer un dictionnaire pour l'accès rapide
    dispo_par_date = {d.date: d for d in disponibilites}
    
    # Récupérer les réservations
    reservations = Reservation.objects.filter(
        maison=maison,
        date_debut__lte=dernier_jour,
        date_fin__gte=premier_jour
    ).exclude(statut='annulee')
    
    # Créer le calendrier
    calendrier = []
    date_courante = premier_jour
    while date_courante <= dernier_jour:
        # Vérifier s'il y a une réservation
        reservation_jour = None
        for reservation in reservations:
            if reservation.date_debut <= date_courante <= reservation.date_fin:
                reservation_jour = reservation
                break
        
        # Récupérer ou créer la disponibilité
        disponibilite = dispo_par_date.get(date_courante)
        if not disponibilite:
            disponibilite = Disponibilite(
                maison=maison,
                date=date_courante,
                disponible=True
            )
        
        calendrier.append({
            'date': date_courante,
            'disponibilite': disponibilite,
            'reservation': reservation_jour,
            'peut_modifier': date_courante >= today
        })
        
        date_courante += timedelta(days=1)
    
    context = {
        'maison': maison,
        'mois': mois,
        'annee': annee,
        'nom_mois': calendar.month_name[mois],
        'calendrier': calendrier,
        'mois_precedent': premier_jour - timedelta(days=1),
        'mois_suivant': dernier_jour + timedelta(days=1),
    }
    
    return render(request, 'reservations/gerer_disponibilites.html', context)


# ======== VUES AJAX ========

@login_required
@require_http_methods(["POST"])
def verifier_disponibilite_ajax(request):
    """Vérifier la disponibilité d'une maison via AJAX"""
    maison_id = request.POST.get('maison_id')
    date_debut = request.POST.get('date_debut')
    date_fin = request.POST.get('date_fin')
    
    if not all([maison_id, date_debut, date_fin]):
        return JsonResponse({'error': 'Paramètres manquants'}, status=400)
    
    try:
        maison = Maison.objects.get(id=maison_id)
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        disponible = Reservation.objects.verifier_disponibilite(maison, date_debut, date_fin)
        
        # Calculer le prix
        nombre_nuits = (date_fin - date_debut).days
        prix_total = maison.prix_par_nuit * nombre_nuits
        
        return JsonResponse({
            'disponible': disponible,
            'nombre_nuits': nombre_nuits,
            'prix_par_nuit': maison.prix_par_nuit,
            'prix_total': prix_total
        })
        
    except (Maison.DoesNotExist, ValueError) as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@gestionnaire_required
@require_http_methods(["POST"])
def modifier_disponibilite_ajax(request):
    """Modifier une disponibilité via AJAX - MODIFIÉ POUR ENTIERS"""
    maison_id = request.POST.get('maison_id')
    date_str = request.POST.get('date')
    disponible = request.POST.get('disponible') == 'true'
    prix_special = request.POST.get('prix_special')
    raison = request.POST.get('raison', '')
    
    try:
        maison = get_object_or_404(Maison, id=maison_id, gestionnaire=request.user)
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Vérifier qu'on ne modifie pas le passé
        if date_obj < timezone.now().date():
            return JsonResponse({'error': 'Impossible de modifier le passé'}, status=400)
        
        # Créer ou mettre à jour la disponibilité - PRIX SPECIAL EN ENTIER
        prix_special_int = None
        if prix_special:
            try:
                prix_special_int = int(float(prix_special))  # Convertir en entier
            except ValueError:
                return JsonResponse({'error': 'Prix spécial invalide'}, status=400)
        
        disponibilite, created = Disponibilite.objects.update_or_create(
            maison=maison,
            date=date_obj,
            defaults={
                'disponible': disponible,
                'prix_special': prix_special_int,
                'raison_indisponibilite': raison if not disponible else ''
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Disponibilité mise à jour pour le {date_obj.strftime("%d/%m/%Y")}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def calculer_prix_ajax(request):
    """Calculer le prix d'une réservation via AJAX - MODIFIÉ POUR ENTIERS"""
    maison_id = request.POST.get('maison_id')
    date_debut = request.POST.get('date_debut')
    date_fin = request.POST.get('date_fin')
    nombre_personnes = int(request.POST.get('nombre_personnes', 1))
    
    try:
        maison = Maison.objects.get(id=maison_id)
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        nombre_nuits = (date_fin - date_debut).days
        
        # Prix de base (en entiers)
        sous_total = maison.prix_par_nuit * nombre_nuits
        
        # Frais de service (exemple: 5% - CALCULÉ EN ENTIERS)
        frais_service = int(sous_total * 0.05)
        
        # Total
        prix_total = sous_total + frais_service
        
        return JsonResponse({
            'nombre_nuits': nombre_nuits,
            'prix_par_nuit': maison.prix_par_nuit,
            'sous_total': sous_total,
            'frais_service': frais_service,
            'prix_total': prix_total,
            'prix_par_personne': prix_total / nombre_personnes
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ======== VUES D'EXPORT ========

@login_required
@gestionnaire_required
def exporter_reservations(request):
    """Exporter les réservations"""
    if request.method == 'POST':
        form = ExportReservationsForm(request.POST)
        if form.is_valid():
            format_export = form.cleaned_data['format_export']
            date_debut = form.cleaned_data['date_debut']
            date_fin = form.cleaned_data['date_fin']
            statuts = form.cleaned_data.get('statuts', [])
            inclure_paiements = form.cleaned_data['inclure_paiements']
            inclure_evaluations = form.cleaned_data['inclure_evaluations']
            
            # Filtrer les réservations
            reservations = Reservation.objects.filter(
                maison__gestionnaire=request.user,
                date_creation__gte=date_debut,
                date_creation__lte=date_fin
            ).select_related('maison', 'client')
            
            if statuts:
                reservations = reservations.filter(statut__in=statuts)
            
            if format_export == 'csv':
                return _export_csv(reservations, inclure_paiements, inclure_evaluations)
            elif format_export == 'excel':
                return _export_excel(reservations, inclure_paiements, inclure_evaluations)
            elif format_export == 'pdf':
                return _export_pdf(reservations, inclure_paiements, inclure_evaluations)
    else:
        form = ExportReservationsForm()
    
    context = {'form': form}
    return render(request, 'reservations/exporter_reservations.html', context)


def _export_csv(reservations, inclure_paiements, inclure_evaluations):
    """Export CSV des réservations - MODIFIÉ POUR ENTIERS"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reservations_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    
    # En-têtes
    headers = [
        'Numéro', 'Client', 'Email Client', 'Maison', 'Date Début', 'Date Fin',
        'Nombre Nuits', 'Nombre Personnes', 'Statut', 'Prix Total', 'Mode Paiement',
        'Date Création'
    ]
    
    if inclure_paiements:
        headers.extend(['Montant Payé', 'Montant Restant'])
    
    if inclure_evaluations:
        headers.extend(['Note Globale', 'Commentaire Évaluation'])
    
    writer.writerow(headers)
    
    # Données
    for reservation in reservations:
        row = [
            reservation.numero,
            reservation.client.get_full_name(),
            reservation.client.email,
            reservation.maison.nom,
            reservation.date_debut,
            reservation.date_fin,
            reservation.nombre_nuits,
            reservation.nombre_personnes,
            reservation.get_statut_display(),
            reservation.prix_total,  # Déjà un entier
            reservation.get_mode_paiement_display(),
            reservation.date_creation.strftime('%Y-%m-%d %H:%M')
        ]
        
        if inclure_paiements:
            montant_paye = reservation.paiements.filter(statut='valide').aggregate(
                total=Sum('montant')
            )['total'] or 0
            row.extend([montant_paye, reservation.prix_total - montant_paye])
        
        if inclure_evaluations:
            try:
                evaluation = reservation.evaluation
                row.extend([evaluation.note_globale, evaluation.commentaire[:100]])
            except:
                row.extend(['', ''])
        
        writer.writerow(row)
    
    return response


def _export_excel(reservations, inclure_paiements, inclure_evaluations):
    """Export Excel des réservations - MODIFIÉ POUR ENTIERS"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Export Excel non disponible", status=400)
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réservations"
    
    # Style des en-têtes
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # En-têtes
    headers = [
        'Numéro', 'Client', 'Email Client', 'Maison', 'Date Début', 'Date Fin',
        'Nombre Nuits', 'Nombre Personnes', 'Statut', 'Prix Total', 'Mode Paiement',
        'Date Création'
    ]
    
    if inclure_paiements:
        headers.extend(['Montant Payé', 'Montant Restant'])
    
    if inclure_evaluations:
        headers.extend(['Note Globale', 'Commentaire Évaluation'])
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Données
    for row_num, reservation in enumerate(reservations, 2):
        ws.cell(row=row_num, column=1, value=reservation.numero)
        ws.cell(row=row_num, column=2, value=reservation.client.get_full_name())
        ws.cell(row=row_num, column=3, value=reservation.client.email)
        ws.cell(row=row_num, column=4, value=reservation.maison.nom)
        ws.cell(row=row_num, column=5, value=reservation.date_debut)
        ws.cell(row=row_num, column=6, value=reservation.date_fin)
        ws.cell(row=row_num, column=7, value=reservation.nombre_nuits)
        ws.cell(row=row_num, column=8, value=reservation.nombre_personnes)
        ws.cell(row=row_num, column=9, value=reservation.get_statut_display())
        ws.cell(row=row_num, column=10, value=reservation.prix_total)  # Entier
        ws.cell(row=row_num, column=11, value=reservation.get_mode_paiement_display())
        ws.cell(row=row_num, column=12, value=reservation.date_creation)
        
        col_offset = 12
        
        if inclure_paiements:
            montant_paye = reservation.paiements.filter(statut='valide').aggregate(
                total=Sum('montant')
            )['total'] or 0
            ws.cell(row=row_num, column=col_offset + 1, value=montant_paye)
            ws.cell(row=row_num, column=col_offset + 2, value=reservation.prix_total - montant_paye)
            col_offset += 2
        
        if inclure_evaluations:
            try:
                evaluation = reservation.evaluation
                ws.cell(row=row_num, column=col_offset + 1, value=evaluation.note_globale)
                ws.cell(row=row_num, column=col_offset + 2, value=evaluation.commentaire[:100])
            except:
                ws.cell(row=row_num, column=col_offset + 1, value="")
                ws.cell(row=row_num, column=col_offset + 2, value="")
    
    # Ajuster les largeurs des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservations_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


def _export_pdf(reservations, inclure_paiements, inclure_evaluations):
    """Export PDF des réservations"""
    # TODO: Implémenter l'export PDF avec ReportLab
    return HttpResponse("Export PDF en cours de développement", status=501)




# Ajouter ces vues dans reservations/views.py

@login_required
@super_admin_required
def gerer_reservation_admin(request, numero):
    """Gestion administrative d'une réservation - SUPER ADMIN UNIQUEMENT"""
    reservation = get_object_or_404(
        Reservation.objects.select_related('maison', 'client'),
        numero=numero
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        motif_admin = request.POST.get('motif_admin', '')
        commentaire_interne = request.POST.get('commentaire_interne', '')
        
        try:
            with transaction.atomic():
                if action == 'valider_admin':
                    # Validation administrative
                    reservation.confirmer(request.user)
                    
                    # Occuper la maison
                    if reservation.maison:
                        reservation.maison.occuper_maison(
                            locataire=reservation.client,
                            date_fin=reservation.date_fin
                        )
                    
                    # Enregistrer les commentaires admin
                    reservation.commentaire_gestionnaire = f"[ADMIN] Validée: {motif_admin}"
                    if commentaire_interne:
                        reservation.commentaire_gestionnaire += f"\nCommentaire interne: {commentaire_interne}"
                    reservation.save()
                    
                    messages.success(request, f'Réservation {reservation.numero} validée par administration.')
                    
                    # TODO: Envoyer notification au client et gestionnaire
                    
                elif action == 'rejeter_admin':
                    # Rejet administratif
                    motif_complet = f"[REJET ADMIN] {motif_admin}"
                    if commentaire_interne:
                        motif_complet += f" - Commentaire interne: {commentaire_interne}"
                    
                    reservation.annuler(motif_complet, request.user)
                    
                    # Libérer la maison si elle était occupée
                    if reservation.maison and reservation.maison.locataire_actuel == reservation.client:
                        reservation.maison.liberer_maison()
                    
                    messages.success(request, f'Réservation {reservation.numero} rejetée par administration.')
                    
                    # TODO: Envoyer notification au client et gestionnaire
                    
                elif action == 'suspendre_admin':
                    # Suspension temporaire
                    reservation.statut = 'suspendue'  # Nécessite d'ajouter ce statut au modèle
                    reservation.commentaire_gestionnaire = f"[ADMIN] Suspendue: {motif_admin}"
                    reservation.save()
                    
                    messages.warning(request, f'Réservation {reservation.numero} suspendue.')
                
                return redirect('reservations:detail', numero=numero)
                
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement administratif: {str(e)}')
    
    # Historique des actions admin sur cette réservation
    historique_admin = []
    if reservation.commentaire_gestionnaire:
        for ligne in reservation.commentaire_gestionnaire.split('\n'):
            if '[ADMIN]' in ligne:
                historique_admin.append(ligne)
    
    context = {
        'reservation': reservation,
        'historique_admin': historique_admin,
        'motifs_validation': [
            'Conformité vérifiée',
            'Documents validés',
            'Paiement confirmé',
            'Validation exceptionnelle',
            'Autre (préciser)'
        ],
        'motifs_rejet': [
            'Documents manquants',
            'Paiement non conforme',
            'Informations erronées',
            'Non-respect des conditions',
            'Fraude suspectée',
            'Autre (préciser)'
        ]
    }
    
    return render(request, 'reservations/gerer_reservation_admin.html', context)


@login_required
@super_admin_required
@require_http_methods(["POST"])
def validation_rapide_admin(request, numero):
    """Validation/rejet rapide via AJAX pour admin"""
    reservation = get_object_or_404(Reservation, numero=numero)
    
    action = request.POST.get('action')
    motif = request.POST.get('motif', 'Action administrative')
    
    try:
        with transaction.atomic():
            if action == 'valider':
                if reservation.statut != 'en_attente':
                    return JsonResponse({'success': False, 'error': 'Réservation non validable'})
                
                reservation.confirmer(request.user)
                
                # Occuper la maison
                if reservation.maison:
                    reservation.maison.occuper_maison(
                        locataire=reservation.client,
                        date_fin=reservation.date_fin
                    )
                
                reservation.commentaire_gestionnaire = f"[ADMIN] Validation rapide: {motif}"
                reservation.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Réservation {reservation.numero} validée',
                    'nouveau_statut': 'confirmee'
                })
                
            elif action == 'rejeter':
                if reservation.statut == 'annulee':
                    return JsonResponse({'success': False, 'error': 'Réservation déjà annulée'})
                
                reservation.annuler(f"[REJET ADMIN] {motif}", request.user)
                
                # Libérer la maison si nécessaire
                if reservation.maison and reservation.maison.locataire_actuel == reservation.client:
                    reservation.maison.liberer_maison()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Réservation {reservation.numero} rejetée',
                    'nouveau_statut': 'annulee'
                })
                
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Action non reconnue'})


@login_required
@super_admin_required
def dashboard_admin_reservations(request):
    """Dashboard spécial pour les administrateurs"""
    # Réservations nécessitant une action admin
    en_attente_validation = Reservation.objects.filter(
        statut='en_attente'
    ).select_related('maison', 'client').order_by('date_creation')
    
    # Réservations suspectes (critères à définir)
    reservations_suspectes = Reservation.objects.filter(
        Q(prix_total__gt=1000000) |  # Montant très élevé
        Q(date_creation__gte=timezone.now() - timedelta(hours=1)) |  # Très récentes
        Q(commentaire_client__icontains='urgent')  # Mentions urgentes
    ).exclude(statut='annulee').select_related('maison', 'client')[:10]
    
    # Statistiques globales
    stats_globales = Reservation.objects.aggregate(
        total_reservations=Count('id'),
        en_attente=Count(Case(When(statut='en_attente', then=1), output_field=IntegerField())),
        confirmees=Count(Case(When(statut='confirmee', then=1), output_field=IntegerField())),
        terminees=Count(Case(When(statut='terminee', then=1), output_field=IntegerField())),
        annulees=Count(Case(When(statut='annulee', then=1), output_field=IntegerField())),
        ca_total=Sum(Case(
            When(statut__in=['confirmee', 'terminee'], then='prix_total'),
            default=0,
            output_field=IntegerField()
        ))
    )
    
    # Actions admin récentes
    reservations_avec_actions_admin = Reservation.objects.filter(
        commentaire_gestionnaire__icontains='[ADMIN]'
    ).select_related('maison', 'client').order_by('-date_modification')[:15]
    
    # Réservations par gestionnaire
    stats_par_gestionnaire = Reservation.objects.filter(
        statut__in=['confirmee', 'terminee']
    ).values(
        'maison__gestionnaire__first_name',
        'maison__gestionnaire__last_name'
    ).annotate(
        total_reservations=Count('id'),
        ca_total=Sum('prix_total')
    ).order_by('-ca_total')[:10]
    
    context = {
        'en_attente_validation': en_attente_validation,
        'reservations_suspectes': reservations_suspectes,
        'stats_globales': stats_globales,
        'reservations_avec_actions_admin': reservations_avec_actions_admin,
        'stats_par_gestionnaire': stats_par_gestionnaire,
        'actions_rapides_disponibles': True
    }
    
    return render(request, 'reservations/dashboard_admin.html', context)


@login_required
@super_admin_required
def historique_actions_admin(request):
    """Historique des actions administratives"""
    # Récupérer toutes les réservations avec actions admin
    reservations_avec_actions = Reservation.objects.filter(
        commentaire_gestionnaire__icontains='[ADMIN]'
    ).select_related('maison', 'client', 'annulee_par').order_by('-date_modification')
    
    # Pagination
    paginator = Paginator(reservations_avec_actions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Filtres
    type_action = request.GET.get('type_action', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    if type_action:
        if type_action == 'validation':
            page_obj.object_list = page_obj.object_list.filter(
                commentaire_gestionnaire__icontains='[ADMIN] Validée'
            )
        elif type_action == 'rejet':
            page_obj.object_list = page_obj.object_list.filter(
                commentaire_gestionnaire__icontains='[REJET ADMIN]'
            )
    
    # Statistiques des actions
    stats_actions = {
        'total_actions': reservations_avec_actions.count(),
        'validations': reservations_avec_actions.filter(
            commentaire_gestionnaire__icontains='[ADMIN] Validée'
        ).count(),
        'rejets': reservations_avec_actions.filter(
            commentaire_gestionnaire__icontains='[REJET ADMIN]'
        ).count(),
    }
    
    context = {
        'page_obj': page_obj,
        'stats_actions': stats_actions,
        'type_action': type_action,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'reservations/historique_actions_admin.html', context)